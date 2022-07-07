import torch
import torch.nn as nn
import torch.optim as optim
from torchvision import transforms as tf
from torch.utils.data import DataLoader

import cv2
import os
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl as xl

import required_classes.config as cf
import required_classes.transform as duke_tf
import required_classes.random_seed as rs
import required_classes.dataset as ds
from models.DAUNet import DAUNet

def visualize_2d(device, model_name_):
    testset = ds.DukeDataset('test',transform = duke_tf.transform_js('test',cf.RESOLUTION))
    testloader = DataLoader(testset, batch_size = cf.BATCH_SIZE,
                                            shuffle=False, num_workers=0, pin_memory=True, drop_last=True)

    # write model's name to "model_name"
    model_name = model_name_
    model_path = os.path.join(cf.MODEL_PATH, model_name)
    modelList = os.listdir(model_path)
    final_model_path = os.path.join(model_path, modelList[-1])
    best_model = torch.load(final_model_path).to(device)
    print(final_model_path)

    modelList = os.listdir(model_path)
    final_model_path = os.path.join(model_path, modelList[-1])
    best_model = torch.load(final_model_path)

    image_result_path = cf.IMAGE_RESULT_PATH + '/' + model_name
    if not os.path.exists(image_result_path): os.mkdir(image_result_path)

    annotation_xl_path = './required_data/Annotation_Boxes_thum.xlsx'

    wb = xl.load_workbook(annotation_xl_path)
    ws = wb.active

    end_slices = [0]
    subject_ids = []


    for i in range(767, 921):
        subject_ids.append(ws.cell(row=i, column=1).value)
        end_slices.append(ws.cell(row=i, column=9).value)

    idx = 0
    with torch.no_grad():
        for batch_idx, data in enumerate(testloader):

            x, y = data
            x = x.to(device)
            y = y.to(device)
            pred = best_model(x)

            x = torch.squeeze(x)
            pred = torch.squeeze(pred)
            y = torch.squeeze(y)

            for i in range(len(pred)):
                j = batch_idx * len(x) + i
                if i%2 == 0:
                    x_1 = x[i].cpu()
                    x_1 = x_1[:3]
                    x_1.transpose_(0, 1)
                    x_1.transpose_(1, 2)
                    
                    x_1= x_1.detach().numpy()
                    y_1 = y[i].cpu().detach().numpy()
                    pred_1 = pred[i]
                    pred_1 = torch.sigmoid(pred_1)
                    pred_1 = (pred_1 > 0.5).type(torch.uint8)
                    pred_1 = pred_1.cpu().detach().numpy()

                    plt.title('{}_{} Image/GT/Pred'.format(subject_ids[idx], j-(end_slices[idx])), fontsize=7)
                    plt.rcParams["figure.figsize"] = (24,8)
                    plt.rcParams['axes.grid'] = True 
                    plt.subplot(1, 3, 1)
                    plt.imshow(x_1)
                    plt.subplot(1, 3, 2)
                    plt.imshow(y_1, cmap='gray')
                    plt.subplot(1, 3, 3)
                    plt.imshow(pred_1, cmap='gray')
                    plt.savefig(image_result_path + '/_{}_{}'.format(subject_ids[idx], j-(end_slices[idx])) + '.png')

                if j == (end_slices[idx+1] - 1):
                    idx += 1