# %%
import torch
import torch.nn as nn
import torch.optim as optim
from torchvision import transforms as tf
from torch.utils.data import DataLoader

import cv2
import re
import os
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl as xl

import required_classes.config as cf
import required_classes.transform as duke_tf
import required_classes.random_seed as rs
import required_classes.dataset as ds
from models.DAUNet import DAUNet

def visualize_3d(device, model_name_):

    testset = ds.DukeDataset('test',transform = duke_tf.transform_js('test',cf.RESOLUTION))
    testloader = DataLoader(testset, batch_size = cf.BATCH_SIZE,
                                            shuffle=False, num_workers=0, pin_memory=True, drop_last=True)

    # write model's name to "model_name"
    model_name = model_name_
    model_path = os.path.join(cf.MODEL_PATH, model_name)
    modelList = os.listdir(model_path)
    final_model_path = os.path.join(model_path, modelList[-1])
    best_model = torch.load(final_model_path).to(device)
    print(final_model_path)

    visualization_path = os.path.join(cf.VISUALIAZATION_PATH, model_name)
    if not os.path.exists(visualization_path): os.mkdir(visualization_path)    
    visualization_path_img = os.path.join(visualization_path, 'img')
    if not os.path.exists(visualization_path_img): os.mkdir(visualization_path_img)
    visualization_path_gt = os.path.join(visualization_path, 'gt')
    if not os.path.exists(visualization_path_gt): os.mkdir(visualization_path_gt)
    visualization_path_pred = os.path.join(visualization_path, 'pred')
    if not os.path.exists(visualization_path_pred): os.mkdir(visualization_path_pred)

    annotation_xl_path = './required_data/Annotation_Boxes_thum.xlsx'

    wb = xl.load_workbook(annotation_xl_path)
    ws = wb.active

    end_slices = [0]
    subject_ids = []


    for i in range(767, 921):
        subject_ids.append(ws.cell(row=i, column=1).value)
        end_slices.append(ws.cell(row=i, column=9).value)

    idx = 0
    with torch.no_grad():
        for batch_idx, data in enumerate(testloader):

            print('---------------- batch_idx {} ----------------'.format(batch_idx))

            x, y = data
            x = x.to(device)
            y = y.to(device)
            pred = best_model(x).to(device)

            x = torch.squeeze(x)
            pred = torch.squeeze(pred)
            y = torch.squeeze(y)

            for i in range(len(pred)):
                j = batch_idx * 20 + i

                x_1 = x[i].cpu()
                x_1 = x_1[3]


                x_1= x_1.detach().numpy()
                y_1 = y[i].cpu().detach().numpy()
                pred_1 = pred[i]
                pred_1 = torch.sigmoid(pred_1)
                pred_1 = (pred_1 > 0.5).type(torch.uint8)
                pred_1 = pred_1.cpu().detach().numpy()


                plt.imsave(os.path.join(visualization_path_img, '{}_{}#{}.png'.format(subject_ids[idx], j-(end_slices[idx]), j)), x_1)
                np.save(os.path.join(visualization_path_gt, '{}_{}#{}.npy'.format(subject_ids[idx], j-(end_slices[idx]), j)), y_1)
                np.save(os.path.join(visualization_path_pred, '{}_{}#{}.npy'.format(subject_ids[idx], j-(end_slices[idx]), j)), pred_1)
                
                if j == (end_slices[idx+1] - 1):
                    idx += 1

    print('---------------Done for saving files---------------')

    # ------------------------3d visualization with matplotlib.pyplot.voxels----------------------
    now = time.localtime()
    str_time = "%02d_%02d_%02d_%02d_%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    print(str_time)

    testset = ds.DukeDataset('test',transform = duke_tf.transform_js('test',cf.RESOLUTION))
    testloader = DataLoader(testset, batch_size = 20,
                                            shuffle=False, num_workers=0, pin_memory=True, drop_last=True)


    # save_path: where 3d_image will be saved
    save_path = os.path.join(cf.IMAGE_RESULT_PATH_3D, model_name)
    if not os.path.exists(save_path): os.mkdir(save_path) 

    list_npy_gt = os.listdir(visualization_path_gt)
    list_npy_pred = os.listdir(visualization_path_pred)
    list_img = os.listdir(visualization_path_img)

    imgs = []
    preds = []
    gts = []

    for i in range(len(list_img)):
        numbers_img = list_img[i].split('#')
        numbers_gt = list_npy_gt[i].split('#')
        numbers_pred = list_npy_pred[i].split('#')

        number_img = int((numbers_img[-1][:-4]))
        number_gt = int((numbers_gt[-1][:-4]))
        number_pred = int((numbers_pred[-1][:-4]))

        imgs.append((list_img[i], number_img))
        preds.append((list_npy_gt[i], number_gt))
        gts.append((list_npy_pred[i], number_pred))


    imgs.sort(key=lambda x: x[-1])
    preds.sort(key=lambda x: x[-1])
    gts.sort(key=lambda x: x[-1])

    wb = xl.load_workbook(annotation_xl_path)
    ws = wb.active

    end_slices = [0]
    subject_ids = []
    total_slices = []
    top_bottom_slices = []
    shapers = []

    for i in range(767, 921):
        subject_ids.append(ws.cell(row=i, column=1).value)
        total_slices.append(ws.cell(row=i, column=10).value)
        end_slices.append(ws.cell(row=i, column=9).value)
        top_bottom_slices.append((int(ws.cell(row=i, column=6).value)-2, ws.cell(row=i, column=7).value))
        shapers.append(int(ws.cell(row=i, column=11).value))
        

    for i in range(len(subject_ids)):
        print('\n----------------' + subject_ids[i] + '----------------\n')
        now = time.localtime()
        str_time = "%02d_%02d_%02d_%02d_%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
        print(str_time)

        img_1s = []
        gt_1s = []
        pred_1s = []
        topbottom_list = range(top_bottom_slices[i][0], top_bottom_slices[i][1]+1)

        idx = 0
        for j in range(end_slices[i], end_slices[i+1]):
            img_1s.append((topbottom_list[idx], cv2.imread(os.path.join(visualization_path_img, imgs[j][0]), cv2.IMREAD_GRAYSCALE)))
            gt_1s.append((topbottom_list[idx], np.load(os.path.join(visualization_path_gt, gts[j][0]))))
            pred_1s.append((topbottom_list[idx], np.load(os.path.join(visualization_path_pred, preds[j][0]))))
            idx += 1

        large_img_1s = []
        large_gt_1s = []
        large_pred_1s = []

        idx = 0
        for k in range(total_slices[i]):
            if k not in topbottom_list:
                large_img_1s.append(np.zeros((256, 256)))
                large_gt_1s.append(np.zeros((256, 256)))
                large_pred_1s.append(np.zeros((256, 256)))
            else:
                if k == img_1s[idx][0]:
                    large_img_1s.append((img_1s[idx][1] - shapers[i]) > 0)
                    large_gt_1s.append(gt_1s[idx][1])
                    large_pred_1s.append(pred_1s[idx][1])
                idx += 1

        print('----start pyplot_voxel----')

        large_img_1s = np.transpose(large_img_1s, (1, 2, 0))
        large_gt_1s = np.transpose(large_gt_1s, (1, 2, 0))
        large_pred_1s = np.transpose(large_pred_1s, (1, 2, 0))

        ax = plt.figure(figsize = (25, 25)).add_subplot(projection='3d')
        ax.voxels(large_img_1s, facecolors='#B3E5FC', edgecolor='none', alpha=.3)
        ax.voxels(large_gt_1s, facecolors='red', edgecolor='none', alpha=.5)
        ax.voxels(large_pred_1s, facecolors='green', edgecolor='none', alpha=.4)


        plt.title(str(subject_ids[i]) + ' / ' + str(top_bottom_slices[i][1] - top_bottom_slices[i][0] + 1) + model_name , fontsize=20)
        plt.savefig(save_path+'/{}.png'.format(str(subject_ids[i])))


# %%
'''
Breast_MRI_766 / 30
Breast_MRI_767 / 47
Breast_MRI_768 / 66
Breast_MRI_769 / 88
Breast_MRI_770 / 101
Breast_MRI_771 / 120
Breast_MRI_772 / 138
Breast_MRI_773 / 155
Breast_MRI_774 / 169
Breast_MRI_775 / 220
Breast_MRI_776 / 333
Breast_MRI_777 / 349
Breast_MRI_778 / 401
Breast_MRI_779 / 415
Breast_MRI_780 / 436
Breast_MRI_781 / 452
Breast_MRI_782 / 470
Breast_MRI_783 / 483
Breast_MRI_784 / 501
Breast_MRI_785 / 514
Breast_MRI_786 / 555
Breast_MRI_787 / 561
Breast_MRI_788 / 571
Breast_MRI_789 / 598
Breast_MRI_790 / 614
Breast_MRI_791 / 627
Breast_MRI_792 / 644
Breast_MRI_793 / 679
Breast_MRI_794 / 706
Breast_MRI_795 / 766
Breast_MRI_796 / 796
Breast_MRI_797 / 815
Breast_MRI_798 / 831
Breast_MRI_799 / 858
Breast_MRI_800 / 875
Breast_MRI_802 / 890
Breast_MRI_803 / 902
Breast_MRI_804 / 955
Breast_MRI_805 / 974
Breast_MRI_806 / 990
Breast_MRI_807 / 1001
Breast_MRI_808 / 1069
Breast_MRI_809 / 1110
Breast_MRI_810 / 1133
Breast_MRI_811 / 1152
Breast_MRI_812 / 1174
Breast_MRI_813 / 1211
Breast_MRI_814 / 1265
Breast_MRI_815 / 1289
Breast_MRI_816 / 1303
Breast_MRI_817 / 1356
Breast_MRI_818 / 1407
Breast_MRI_819 / 1433
Breast_MRI_820 / 1517
Breast_MRI_821 / 1523
Breast_MRI_822 / 1559
Breast_MRI_823 / 1588
Breast_MRI_824 / 1611
Breast_MRI_825 / 1623
Breast_MRI_826 / 1644
Breast_MRI_827 / 1657
Breast_MRI_828 / 1669
Breast_MRI_829 / 1688
Breast_MRI_830 / 1719
Breast_MRI_831 / 1787
Breast_MRI_832 / 1811
Breast_MRI_833 / 1831
Breast_MRI_834 / 1853
Breast_MRI_835 / 1870
Breast_MRI_836 / 1878
Breast_MRI_837 / 1893
Breast_MRI_838 / 1913
Breast_MRI_839 / 1959
Breast_MRI_840 / 1986
Breast_MRI_841 / 2003
Breast_MRI_842 / 2023
Breast_MRI_843 / 2043
Breast_MRI_844 / 2057
Breast_MRI_845 / 2077
Breast_MRI_846 / 2087
Breast_MRI_847 / 2104
Breast_MRI_848 / 2159
Breast_MRI_849 / 2177
Breast_MRI_850 / 2230
Breast_MRI_851 / 2254
Breast_MRI_852 / 2289
Breast_MRI_853 / 2312
Breast_MRI_854 / 2353
Breast_MRI_855 / 2358
Breast_MRI_856 / 2372
Breast_MRI_857 / 2413
Breast_MRI_858 / 2424
Breast_MRI_859 / 2447
Breast_MRI_860 / 2480
Breast_MRI_861 / 2514
Breast_MRI_862 / 2535
Breast_MRI_863 / 2562
Breast_MRI_864 / 2585
Breast_MRI_865 / 2610
Breast_MRI_866 / 2637
Breast_MRI_867 / 2661
Breast_MRI_868 / 2689
Breast_MRI_869 / 2726
Breast_MRI_870 / 2735
Breast_MRI_871 / 2747
Breast_MRI_872 / 2756
Breast_MRI_873 / 2790
Breast_MRI_874 / 2887
Breast_MRI_875 / 2912
Breast_MRI_876 / 2918
Breast_MRI_877 / 2935
Breast_MRI_878 / 2955
Breast_MRI_879 / 3011
Breast_MRI_880 / 3037
Breast_MRI_881 / 3051
Breast_MRI_882 / 3081
Breast_MRI_883 / 3103
Breast_MRI_884 / 3137
Breast_MRI_885 / 3158
Breast_MRI_886 / 3187
Breast_MRI_887 / 3222
Breast_MRI_888 / 3235
Breast_MRI_889 / 3240
Breast_MRI_890 / 3254
Breast_MRI_891 / 3271
Breast_MRI_892 / 3283
Breast_MRI_894 / 3297
Breast_MRI_895 / 3322
Breast_MRI_896 / 3359
Breast_MRI_897 / 3368
Breast_MRI_898 / 3387
Breast_MRI_899 / 3418
Breast_MRI_900 / 3441
Breast_MRI_901 / 3470
Breast_MRI_902 / 3481
Breast_MRI_903 / 3498
Breast_MRI_904 / 3546
Breast_MRI_905 / 3568
Breast_MRI_906 / 3581
Breast_MRI_907 / 3629
Breast_MRI_908 / 3648
Breast_MRI_909 / 3662
Breast_MRI_910 / 3711
Breast_MRI_911 / 3728
Breast_MRI_912 / 3744
Breast_MRI_913 / 3755
Breast_MRI_914 / 3790
Breast_MRI_915 / 3823
Breast_MRI_916 / 3839
Breast_MRI_917 / 3922
Breast_MRI_918 / 3948
Breast_MRI_919 / 3962
Breast_MRI_920 / 3979
Breast_MRI_921 / 4000
'''
