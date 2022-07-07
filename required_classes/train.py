# %%
from pyparsing import col
import torch
import torch.nn as nn
import torch.optim as optim
from torchvision import transforms as tf
from torch.utils.data import DataLoader

import sys
import os
import stat
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl as xl

import required_classes.config as cf
import required_classes.transform as duke_tf
import required_classes.random_seed as rs
import required_classes.dataset as ds
from required_classes.metrics.total_score import total_score
from models.DAUNet import DAUNet
from required_classes.metrics.dsc_score import dice_score
from required_classes.loss_functions.customLoss_dice_bce import customLoss
from required_classes.loss_functions.customLoss_dice_bce3 import customLoss as customLoss3
from required_classes.loss_functions.customLoss_dice_bce3_exp import customLoss as customLoss3_exp
from required_classes.config import parse_args_

class Train_DBC:

    def __init__(self, args):
        self.args = args

    def __call__(self):
        self.run()

    def run(self):
        n_args = self.args
        
        trainset = ds.DukeDataset('train',transform = duke_tf.transform_js('train',cf.RESOLUTION))
        trainloader = DataLoader(trainset, batch_size = cf.BATCH_SIZE,
                                                shuffle=True, num_workers=0, pin_memory=True, drop_last=True) #num_workers=0으로 놓은 이유는 오류를 방지하기 위해서 & 원래 default값이 0임..!

        testset = ds.DukeDataset('test',transform = duke_tf.transform_js('test',cf.RESOLUTION))
        testloader = DataLoader(testset, batch_size = cf.BATCH_SIZE,
                                                shuffle=False, num_workers=0, pin_memory=True, drop_last=True)

        model = DAUNet(cf.STARTING_CHANNEL, cf.ENCODING_CHANNELS, cf.DECODING_CHANNELS, cf.GROWTH_RATE).to(n_args['device'])
        creterion = customLoss3_exp
        optimizer = optim.AdamW(model.parameters(), lr = 0)
        lr_sche = optim.lr_scheduler.OneCycleLR(optimizer, max_lr = cf.MAX_LR, epochs=cf.NUM_EPOCHS, steps_per_epoch=len(trainloader))

        #get latest model number from log/latest_model_number.txt
        f = open(cf.LOG_MODEL_NUMBER, 'r')
        latest_model_number = f.read()
        f.close()

        #set model name
        model_name = cf.MODEL_NAME + '_' +latest_model_number
        print('---model_name: ',model_name,'---')

        now = time.localtime()
        str_time = "%02d_%02d_%02d_%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour)
        log_path = os.path.join(cf.TOTAL_RESULT_PATH, model_name + '.txt')


        if n_args['log_param']:

            # create total result txt file
            f = open(log_path, 'w')
            f.close()

            # update model number at log/latest_model_number.txt
            f = open(cf.LOG_MODEL_NUMBER, 'w')
            f.write(str(int(latest_model_number) + 1))
            f.close()

            # make paths and directory for models 
            model_path = cf.MODEL_PATH + '/' + model_name
            os.mkdir(model_path)

        keys = ['Start Time', 'Model_name', 'Optimizer', 'LR_scheduler', 'LOSS',
                'Growth Rate', 'Starting Channel', 'Encoding Channels', 'Decoding Channels',
                'Random seed', 'Resolution', 'Batch_size', 'Num_Epochs', 'Max_LR', 'Smooth',
                'POS_weight', 'BCE_weight', 'FP_weight', 'Dice_weight',
            ]
        values = [str_time, cf.MODEL_NAME, cf.OPTIMIZER, cf.LR_SCHEDULER, cf.LOSS, 
                    str(cf.GROWTH_RATE), str(cf.STARTING_CHANNEL), '/'.join(map(str, cf.ENCODING_CHANNELS)), '/'.join(map(str, cf.DECODING_CHANNELS)), 
                    str(cf.RANDOM_SEED), str(cf.RESOLUTION), str(cf.BATCH_SIZE), str(cf.NUM_EPOCHS), str(cf.MAX_LR), str(cf.SMOOTH), 
                    str(cf.POS_WEIGHT), str(cf.BCE_WEIGHT), str(cf.BCE2_WEIGHT), str(cf.DSC_WEIGHT),
            ]
        f = open(log_path, 'a')
        for i in range(len(keys)):
            f.write(keys[i] + ' ' + values[i] + '\n')
            print(keys[i], ' ', values[i])

        f.close()

        st = os.stat('./output')
        os.chmod('./output', st.st_mode | stat.S_IEXEC)

        if not os.path.exists(cf.OUTPUT_PATH): os.mkdir(cf.OUTPUT_PATH)    
        if not os.path.exists(cf.MODEL_PATH): os.mkdir(cf.MODEL_PATH)    
        if not os.path.exists(cf.IMAGE_RESULT_PATH): os.mkdir(cf.IMAGE_RESULT_PATH)   
        if not os.path.exists(cf.IMAGE_RESULT_PATH_3D): os.mkdir(cf.IMAGE_RESULT_PATH_3D) 
        if not os.path.exists(cf.TOTAL_RESULT_PATH): os.mkdir(cf.TOTAL_RESULT_PATH)    
        if not os.path.exists(cf.GRAPH_PATH): os.mkdir(cf.GRAPH_PATH)    
        if not os.path.exists(cf.VISUALIAZATION_PATH): os.mkdir(cf.VISUALIAZATION_PATH)    

        rs.Fix_randomness(cf.RANDOM_SEED)
        print('Decoding channel : ',cf.DECODING_CHANNELS)

        # %%
        def train_model(_model, _epoch, _train_dataset, _train_dataloader, _optimizer, _lr_sche, _creterion, _array_cost, _array_accu_zero, _array_accu_nonzero, _array_cost_bce, _array_cost_bce2, _array_cost_dsc, _log_path):
            now = time.localtime()
            str_time = "%02d_%02d_%02d_%02d_%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
            print('Traing Epoch: {} / Time: {}'.format(_epoch, str_time))
            
            train_cost_sum = 0
            bce_cost_sum = 0
            bce2_cost_sum = 0
            dsc_cost_sum = 0
            dice_score_sum = 0
            zero_dice_score_sum = 0
            zero_count_sum = 0
            nonzero_count_sum = 0
            print('----------------train----------------')
            for batch_idx, data in enumerate(_train_dataloader):
                x, y = data
                x = x.to(n_args['device'])
                y = y.to(n_args['device'])

                _optimizer.zero_grad()
                prediction = _model(x)

                zero_dsc, nonzero_dsc, zero_count, nonzero_count = dice_score(prediction, y)
                zero_dice_score_sum += zero_dsc
                dice_score_sum += nonzero_dsc
                zero_count_sum += zero_count
                nonzero_count_sum += nonzero_count

                cost_bce2, cost_bce, cost_dsc = _creterion(prediction, y, cf.POS_WEIGHT, n_args['device'], cf.SMOOTH)
                
                bce2_cost_sum += cost_bce2.item()
                bce_cost_sum += cost_bce.item() * cf.BCE_WEIGHT
                dsc_cost_sum += cost_dsc.item() * cf.DSC_WEIGHT

                cost = cost_bce2*torch.Tensor([cf.BCE2_WEIGHT]).to(n_args['device']) + cost_bce*torch.Tensor([cf.BCE_WEIGHT]).to(n_args['device']) + cost_dsc*torch.Tensor([cf.DSC_WEIGHT]).to(n_args['device'])
                cost.backward()
                _optimizer.step()
                _lr_sche.step()

                train_cost_sum += cost.item()
                # if (batch_idx + 1) % (len(_train_dataloader)//2) == 0
                if (batch_idx + 1) % (20) == 0:
                    train_cost_avg = train_cost_sum / (batch_idx + 1)
                    train_cost_bce2_avg = bce2_cost_sum / (batch_idx + 1)
                    train_cost_bce_avg = bce_cost_sum / (batch_idx + 1)
                    train_cost_dsc_avg = dsc_cost_sum / (batch_idx + 1)
                    log = 'Epoch:{:2d}/{} batch_idx:{:3d} Train Cost Total:{:2.6f} BCE:{:2.6f} NEW_LOSS:{:2.6f} DSC:{:2.6f}'.format(_epoch, cf.NUM_EPOCHS, batch_idx+1, train_cost_avg, train_cost_bce_avg, train_cost_bce2_avg, train_cost_dsc_avg)
                    print(log)
                    if n_args['log_param']:    
                        f = open(_log_path, 'a')
                        f.write(log + '\n')
                        f.close()
                
            nonzero_accuracy = (dice_score_sum / nonzero_count_sum)
            zero_accuracy = (zero_dice_score_sum/ zero_count_sum)
            log = 'Epoch:{:2d}/{} batch_idx:{:3d} Train Accuracy:zero{:2.6f}_nonzero:{:2.6f}'.format(_epoch, cf.NUM_EPOCHS, batch_idx, zero_accuracy, nonzero_accuracy)
            print(log)
            if n_args['log_param']:    
                f = open(_log_path, 'a')
                f.write('\nTrain Accuracy:\n    '+ log + '\n')
                f.close()
                _array_cost.append(train_cost_sum/len(_train_dataloader))
                _array_cost_bce.append(bce_cost_sum/len(_train_dataloader))
                _array_cost_bce2.append(bce2_cost_sum/len(_train_dataloader))
                _array_cost_dsc.append(dsc_cost_sum/len(_train_dataloader))
                _array_accu_zero.append(zero_accuracy)
                _array_accu_nonzero.append(nonzero_accuracy)

        # %%
        def test_model(_model, _test_dataset, _test_dataloader, _creterion, _array_cost, _array_accu_zero, _array_accu_nonzero, _array_cost_bce, _array_cost_bce2, _array_cost_dsc, _log_path):
            test_cost_sum = 0
            bce_cost_sum = 0
            bce2_cost_sum = 0
            dsc_cost_sum = 0
            dice_score_sum = 0
            zero_dice_score_sum = 0
            zero_count_sum = 0
            nonzero_count_sum = 0
            print('----------------test----------------')
            with torch.no_grad():
                for batch_idx, data in enumerate(_test_dataloader):
                    x, y = data
                    x = x.to(n_args['device'])
                    y = y.to(n_args['device'])
                    prediction = _model(x)
                        
                    zero_dsc, nonzero_dsc, zero_count, nonzero_count = dice_score(prediction, y)
                    zero_dice_score_sum += zero_dsc
                    dice_score_sum += nonzero_dsc
                    zero_count_sum += zero_count
                    nonzero_count_sum += nonzero_count

                    cost_bce2, cost_bce, cost_dsc = _creterion(prediction, y, cf.POS_WEIGHT, n_args['device'], cf.SMOOTH)
                    bce2_cost_sum += cost_bce2.item() * cf.BCE2_WEIGHT 
                    bce_cost_sum += cost_bce.item() * cf.BCE_WEIGHT 
                    dsc_cost_sum += cost_dsc.item() * cf.DSC_WEIGHT

                    test_cost_sum += ( cost_bce.item() * cf.BCE_WEIGHT + cost_bce2.item() * cf.BCE2_WEIGHT + cost_dsc.item() * cf.DSC_WEIGHT )

            test_cost_avg = test_cost_sum/len(_test_dataloader)
            test_cost_bce_avg = bce_cost_sum / len(_test_dataloader)
            test_cost_bce2_avg = bce2_cost_sum / len(_test_dataloader)
            test_cost_dsc_avg = dsc_cost_sum / len(_test_dataloader)
            zero_accuracy = (zero_dice_score_sum/ zero_count_sum)    
            nonzero_accuracy = (dice_score_sum / nonzero_count_sum)
            log = 'Test Cost ToTal: {:2.6f} BCE: {:2.6f} NEW_LOSS: {:2.6f} DSC {:2.6f} Test Accuracy:zero{:2.6f}_nonzero:{:2.6f}'.format(test_cost_avg, test_cost_bce_avg, test_cost_bce2_avg, test_cost_dsc_avg, zero_accuracy, nonzero_accuracy)
            print(log)
            if n_args['log_param']:    
                f = open(_log_path, 'a')
                f.write('Test Accuracy:\n    '+ log + '\n\n')
                f.close()
                _array_cost.append(test_cost_avg)
                _array_cost_bce.append(test_cost_bce_avg)
                _array_cost_bce2.append(test_cost_bce2_avg)
                _array_cost_dsc.append(test_cost_dsc_avg)
                _array_accu_zero.append(zero_accuracy)
                _array_accu_nonzero.append(nonzero_accuracy)

        # %%
        def test_model_total_score(_testloader, model_name, _log_path):
            print('------------------start calculating total score all the results are saved------------------')
            classification_sum = 0
            dsc_sum = 0
            iou_sum = 0
            pixaccu_sum = 0
            precision_sum = 0
            recall_sum = 0
            hd_sum = 0
            zero_count = 0
            nonzero_count = 0

            model_path = os.path.join(cf.MODEL_PATH, model_name)
            modelList = os.listdir(model_path)
            final_model_path = os.path.join(model_path, modelList[-1])
            best_model = torch.load(final_model_path).to(n_args['device'])
            print(final_model_path)

            with torch.no_grad():
                for batch_idx, data in enumerate(_testloader):

                    x, y = data
                    x = x.to(n_args['device'])
                    y = y.to(n_args['device'])
                    pred = best_model(x).to(n_args['device'])

                    classification_sc, dsc_sc, iou_sc, pixaccu_sc, precision_sc, recall_sc, hd_sc, zero_c, nonzero_c = total_score(pred, y)
                    classification_sum += classification_sc
                    dsc_sum += dsc_sc
                    iou_sum += iou_sc
                    pixaccu_sum += pixaccu_sc
                    recall_sum += recall_sc
                    precision_sum += precision_sc
                    hd_sum += hd_sc
                    zero_count += zero_c
                    nonzero_count += nonzero_c

                    
                    if (batch_idx+1)%10 == 0:
                        print('-----batch_idx-----:', batch_idx+1)
                        infos = 'Accuracys:\n    zero_slice_classification:', classification_sum/zero_count , '\n    dsc:',dsc_sum/nonzero_count, '/ iou:',iou_sum/nonzero_count, '/ pixcel accuracy:', pixaccu_sum/nonzero_count, '\n    precision:', precision_sum/nonzero_count, '/ recall:', recall_sum/nonzero_count, '/ housdorff distance:' , hd_sum/nonzero_count
                        infos = ' '.join(map(str, infos))
                        print(infos)
                        info_list = [('zero_slice_classification', classification_sum/zero_count),
                                     ('dsc',dsc_sum/nonzero_count), 
                                     ('iou',iou_sum/nonzero_count),
                                     ('pixcel accuracy', pixaccu_sum/nonzero_count),
                                     ('precision', precision_sum/nonzero_count),
                                     ('recall', recall_sum/nonzero_count),
                                     ('housdorff distance' , hd_sum/nonzero_count) ]
                    
            
            print("\n\n-----------------------------Total Average result-----------------------------\n\n")
            print(infos)

            if n_args['log_param']:
                f = open(_log_path, 'a')
                f.write("\n\n-----------------------------Total Average result-----------------------------\n\n" + infos)
                f.close()

            return info_list

        # %%
        #RecordTime

        train_cost_arr = []
        train_cost_bce_arr = []
        train_cost_bce2_arr = []
        train_cost_dsc_arr = []
        train_zero_dsc_arr = []
        train_nonzero_dsc_arr = []

        test_cost_arr = []
        test_cost_bce_arr = []
        test_cost_bce2_arr = []
        test_cost_dsc_arr = []
        test_zero_dsc_arr = []
        test_nonzero_dsc_arr = []

        x_array = []
        str_encblocks = '|'.join(map(str, cf.ENCODING_CHANNELS)) #for saving model informaiton
        str_decblocks = '|'.join(map(str, cf.DECODING_CHANNELS)) #for saving model informaiton

        for epoch in range(cf.NUM_EPOCHS):
            x_array.append(epoch + 1)
            train_cost_sum = 0
            test_cost_sum = 0

            model.train()
            train_model(model, epoch+1, trainset, trainloader, optimizer, lr_sche, creterion, train_cost_arr, train_zero_dsc_arr, train_nonzero_dsc_arr, train_cost_bce_arr, train_cost_bce2_arr, train_cost_dsc_arr, log_path)

            model.eval()
            test_model(model, testset, testloader, creterion, test_cost_arr, test_zero_dsc_arr, test_nonzero_dsc_arr, test_cost_bce_arr, test_cost_bce2_arr, test_cost_dsc_arr, log_path)
        
            '''If you use codes below you can see Train and Test COST & ACCURACY per each epoch'''
            plt.rcParams["figure.figsize"] = (21,7)
            plt.rcParams['axes.grid'] = True     
            
            plt.subplot(1,3,1)
            plt.plot(x_array, train_cost_arr, 'r', label='train cost')
            plt.plot(x_array, test_cost_arr, 'g', label='test cost')
            plt.xlabel('EPOCH')
            plt.ylabel('Cost')
            plt.title('Train/Test Cost per Epoch')
            
            plt.subplot(1,3,2)
            plt.plot(x_array, train_zero_dsc_arr, 'r', label='train accu')
            plt.plot(x_array, test_zero_dsc_arr, 'g', label='test accu')
            plt.xlabel('EPOCH')
            plt.ylabel('ZERO ACCURACY')
            plt.title('{}Train/Test Zero_ACCURACY per Epoch{}'.format(model_name, epoch))
            
            plt.subplot(1,3,3)
            plt.plot(x_array, train_nonzero_dsc_arr, 'r', label='train accu')
            plt.plot(x_array, test_nonzero_dsc_arr, 'g', label='test accu')
            plt.xlabel('EPOCH')
            plt.ylabel('NONZERO ACCURACY')
            plt.title('Train/Test Nonzero_ACCURACY per Epoch')
            
            if n_args['log_param']:
                plt.savefig(os.path.join(cf.GRAPH_PATH, model_name) + '.png' ,facecolor='#ffffff',edgecolor='blue')
            plt.show()

            if test_nonzero_dsc_arr[-1] > 0.4 and max(test_nonzero_dsc_arr) == test_nonzero_dsc_arr[-1]:
                torch.save(model, model_path + '/{}_ep:{}_{:1.5f}.pth'.format(model_name, epoch ,max(test_nonzero_dsc_arr)))
                print(model_path + '{}_{:1.3f}.pth'.format(epoch ,max(test_nonzero_dsc_arr)))

        infos = test_model_total_score(testloader, model_name, log_path)

        # %%
        #Openpyxl setting
        if n_args['log_param']:
            wb = xl.load_workbook(cf.XL_PATH)  #엑셀 파일을 불러오는 과정
            wb.create_sheet(model_name, 0)
            wb.save(cf.XL_PATH)

            ws = wb[model_name] #생성한 sheet를 선택

            for i in range(len(keys)):
                ws.cell(row=i+1, column=1).value = keys[i]
                ws.cell(row=i+1, column=3).value = values[i]

            wb.save(cf.XL_PATH)    
            start_row = len(keys) + 4

            columns = ['Epoch', 'train_cost', 'train_bce_cost', 'train_dsc_cost', 'test_zero_dsc',  'test_nonzero_dsc']
            for i in range(len(columns)):
                ws.cell(row=start_row-1, column=(2*i+1)).value=columns[i]

            for i in range(len(x_array)):
                ws.cell(row=i+start_row, column=1).value = x_array[i]
                ws.cell(row=i+start_row, column=3).value = train_cost_arr[i]
                ws.cell(row=i+start_row, column=5).value = train_cost_bce_arr[i]
                ws.cell(row=i+start_row, column=7).value = train_cost_dsc_arr[i]
                ws.cell(row=i+start_row, column=9).value = test_zero_dsc_arr[i]
                ws.cell(row=i+start_row, column=11).value = test_nonzero_dsc_arr[i]

            for i in range(len(infos)):
                ws.cell(row = start_row + len(x_array) + 2 + i, column=1).value = infos[i][0]
                ws.cell(row = start_row + len(x_array) + 2 + i, column=3).value = infos[i][1]

            wb.save(cf.XL_PATH)   

        return model_name